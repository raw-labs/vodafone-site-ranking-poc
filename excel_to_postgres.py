import openpyxl
import json
import sys
import re
import os
import hashlib
from datetime import datetime
from openpyxl.worksheet.worksheet import Worksheet

MAX_IDENTIFIER_LENGTH = 63  # Postgres limit for identifiers

def finalize_column_name(raw_name, used_names):
    """
    Convert 'raw_name' into a safe, unique PostgreSQL identifier:
      1) sanitize => underscores
      2) if >63 chars => short-hash approach
      3) if still collision => append _2, _3, etc., re-check length
      4) return final name, add to used_names
    """
    name = raw_name.lower()
    # Replace any non-alphanumeric with _
    name = re.sub(r'\W+', '_', name)
    # Strip leading/trailing underscores
    name = name.strip('_')
    if not name:
        name = "col"

    # If name > 63 chars => shorten with hash
    if len(name) > MAX_IDENTIFIER_LENGTH:
        name = shorten_with_hash(name)

    # Deduplicate if already used
    base = name
    counter = 2
    while name in used_names:
        candidate = f"{base}_{counter}"
        counter += 1
        if len(candidate) > MAX_IDENTIFIER_LENGTH:
            candidate = shorten_with_hash(candidate)
        name = candidate

    used_names.add(name)
    return name

def shorten_with_hash(long_name):
    """
    Keep first 50 chars + '_' + 8-char md5 hash
    Then if still >63, slice to 63.
    """
    shortened = long_name[:50]
    h = hashlib.md5(long_name.encode('utf-8')).hexdigest()[:8]
    new_name = f"{shortened}_{h}"
    if len(new_name) > MAX_IDENTIFIER_LENGTH:
        new_name = new_name[:MAX_IDENTIFIER_LENGTH]
    return new_name

def load_config(config_path):
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def parse_region(region_str):
    """
    region_str like 'B2:F100'
    parse -> (start_col, start_row, end_col, end_row) zero-based
    if invalid or None => return None to read entire used range
    """
    if not region_str:
        return None

    pattern = re.compile(r'^([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)$')
    match = pattern.match(region_str)
    if not match:
        print(f"Warning: region '{region_str}' not recognized. Using full sheet.")
        return None

    start_col_letters, start_row_num, end_col_letters, end_row_num = match.groups()
    start_col = letters_to_index(start_col_letters)
    start_row = int(start_row_num) - 1
    end_col = letters_to_index(end_col_letters)
    end_row = int(end_row_num) - 1
    return (start_col, start_row, end_col, end_row)

def letters_to_index(col_letters):
    """ e.g. 'A' -> 0, 'B'->1, 'AA'->26, etc. """
    col_letters = col_letters.upper()
    result = 0
    for c in col_letters:
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result - 1

def unmerge_worksheet(ws: Worksheet):
    """
    For every merged cell range, copy the top-left cell’s value into all cells 
    within that range, then unmerge it so each cell can be read or written 
    individually without triggering the read-only 'MergedCell' error.
    """
    merged_ranges = list(ws.merged_cells.ranges)  # snapshot of ranges
    for merged_range in merged_ranges:
        # merged_range.bounds -> (min_col, min_row, max_col, max_row)
        min_col, min_row, max_col, max_row = merged_range.bounds

        # The top-left cell is (min_row, min_col) in openpyxl's row-col format
        top_left_value = ws.cell(row=min_row, column=min_col).value

        # First unmerge the cells in this range
        ws.unmerge_cells(str(merged_range))

        # Now fill every cell in that former merged region with the top-left value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c).value = top_left_value


def generate_ddl(table_name, column_names, approach='stringAll'):
    """
    Create a CREATE TABLE statement for Postgres.
      column_names: [(col_name, sample_val), ...]
      approach => 'stringAll' or 'auto'
    """
    col_defs = []
    for (col_name, sample_val) in column_names:
        if approach == 'stringAll':
            col_type = 'VARCHAR(512)'
        else:
            col_type = guess_type(sample_val)
        col_defs.append(f'  "{col_name}" {col_type}')
    col_defs_str = ",\n".join(col_defs)
    ddl = f'CREATE TABLE "{table_name}" (\n{col_defs_str}\n);\n'
    return ddl

def guess_type(sample_val):
    """
    If sample_val is None => text
    If int/float => numeric
    else => try float parse => numeric, else text
    """
    if sample_val is None:
        return "VARCHAR(512)"
    if isinstance(sample_val, (int, float)):
        return "NUMERIC"
    # attempt float parse
    val_str = str(sample_val)
    try:
        float(val_str)
        return "NUMERIC"
    except:
        pass
    return "VARCHAR(512)"

def convert_value(raw_val, approach):
    """
    If stringAll => always store as text
    If auto => numeric if possible, else text
    """
    if raw_val is None:
        return "NULL"
    if approach == "stringAll":
        escaped = str(raw_val).replace("'", "''")
        return f"'{escaped}'"
    else:
        # auto approach
        if isinstance(raw_val, (int, float)):
            return str(raw_val)
        if isinstance(raw_val, datetime):
            text_val = str(raw_val)
            esc = text_val.replace("'", "''")
            return f"'{esc}'"
        # try numeric parse
        text_val = str(raw_val)
        try:
            float(text_val)
            return text_val
        except:
            pass
        esc = text_val.replace("'", "''")
        return f"'{esc}'"

def read_excel_content(excel_file, config, extra_col_name=None, extra_col_value=None):
    """
    Loads the Excel file, unmerges (filling each cell),
    reads the specified sheet, identifies headers, and returns:
      (headerList, dataRows, colSamples)

    The "region" from config (if any) bounds the reading.
    'headerRows' is how many rows define the column names.
    We combine them to produce a single name per column.
    This version uses name-based alignment for subsequent merges:

    - We'll produce a dictionary: finalColName -> columnIndex
      for the CURRENT file, then return data as a list of dicts
      so that missing columns can be recognized later.

    If extra_col_name/extra_col_value is set, every data dict will
    include { extra_col_name: extra_col_value }.
    """
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws_name = config.get("sheetName", "Sheet1")
    if ws_name not in wb.sheetnames:
        # Return None to signal "skip this file"
        print(f"Warning: File '{excel_file}' has no sheet '{ws_name}'; skipping.")
        return None, None, None

    ws = wb[ws_name]

    # unmerge all cells so each has correct value
    unmerge_worksheet(ws)

    region_tuple = parse_region(config.get("region"))
    approach = config.get("typeApproach", "stringAll")
    header_rows = config.get("headerRows", 1)
    header_joiner = config.get("headerJoiner", " - ")

    # figure out row/col bounds
    # We'll read the entire region, but interpret columns by header text
    if region_tuple:
        sc, sr, ec, er = region_tuple
    else:
        sc = ws.min_column - 1
        sr = ws.min_row - 1
        ec = ws.max_column - 1
        er = ws.max_row - 1

    # read all rows in that region
    all_rows = []
    for row_idx in range(sr, er + 1):
        row_cells = []
        for col_idx in range(sc, ec + 1):
            cell_val = ws.cell(row=row_idx + 1, column=col_idx + 1).value
            row_cells.append(cell_val)
        all_rows.append(row_cells)

    if len(all_rows) < header_rows:
        print(f"Warning: Not enough rows for 'headerRows' in file {excel_file}, skipping.")
        return None, None, None

    # Build combined header from the top 'headerRows'
    num_cols = len(all_rows[0])
    header_labels = []
    for col_i in range(num_cols):
        parts = []
        for hr_i in range(header_rows):
            val = all_rows[hr_i][col_i]
            if val is None or str(val).strip() == "":
                val = f"col{hr_i+1}_c{col_i+1}"
            else:
                val = str(val).strip()
            parts.append(val)
        combined_name = header_joiner.join(parts)
        header_labels.append(combined_name)

    # Deduplicate/truncate them => produce final col names for THIS file
    used = set()
    final_header_for_file = []
    for raw_col_name in header_labels:
        safe_name = finalize_column_name(raw_col_name, used)
        final_header_for_file.append(safe_name)

    # Data rows => skip the 'headerRows'
    data_part = all_rows[header_rows:]

    # Convert data_part to a list of dictionaries: {colName: cellValue}
    # (by position for this single file)
    all_dicts = []
    for row in data_part:
        row_dict = {}
        for c_idx, raw_val in enumerate(row):
            col_name = final_header_for_file[c_idx]
            row_dict[col_name] = raw_val
        # If extra col was requested, add it
        if extra_col_name:
            if extra_col_name not in final_header_for_file:
                final_header_for_file.append(extra_col_name)
            row_dict[extra_col_name] = extra_col_value
        all_dicts.append(row_dict)

    # For auto approach => find sample for each column in this file
    col_samples = {}
    if approach == "auto":
        for row_dict in all_dicts:
            for col_name, val in row_dict.items():
                if col_name not in col_samples and val is not None:
                    col_samples[col_name] = val

    return final_header_for_file, all_dicts, col_samples

def single_excel_to_sql_inserts(excel_file, config):
    """
    Processes a single Excel file -> returns (ddl, [insert statements]).
    The script uses name-based reading but there's only 1 file, so there's
    no "merging of columns" from multiple sources.
    """
    approach = config.get("typeApproach", "stringAll")
    table_name = config["tableName"]

    final_header_for_file, all_dicts, col_samples_dict = read_excel_content(excel_file, config)
    if final_header_for_file is None:
        # means "skip" => just return empty
        return f"-- Skipped file {excel_file}, no sheet or not enough rows.\n", []

    # If approach == auto => pick sample from col_samples_dict
    # We'll produce columns_info in the order of final_header_for_file
    if col_samples_dict is None:
        col_samples_dict = {}

    columns_info = []
    for col_name in final_header_for_file:
        sample_val = col_samples_dict.get(col_name, None) if approach == "auto" else None
        columns_info.append((col_name, sample_val))

    ddl = generate_ddl(table_name, columns_info, approach)

    # build insert statements
    insert_statements = []
    for row_dict in all_dicts:
        col_list = []
        val_list = []
        for col_name, sample_val in columns_info:
            raw_val = row_dict.get(col_name)
            val_str = convert_value(raw_val, approach)
            col_list.append(f'"{col_name}"')
            val_list.append(val_str)
        ins_sql = f'INSERT INTO "{table_name}" ({", ".join(col_list)}) VALUES ({", ".join(val_list)});'
        insert_statements.append(ins_sql)

    return ddl, insert_statements

def multi_excel_to_sql_inserts(config):
    """
    Scans a directory for Excel files, merges them into one table.
    Variation: If a file has no 'sheetName', skip it.
               If columns are missing, fill with NULL.
               If columns are extra, ignore them.
    """
    directory = config["directory"]
    pattern = config.get("filenamePattern")
    filename_col = config.get("filenameColumnName")
    approach = config.get("typeApproach", "stringAll")
    table_name = config["tableName"]

    if not os.path.isdir(directory):
        raise ValueError(f"Directory '{directory}' not found or not a directory.")

    # Compile filename pattern if any
    rx = re.compile(pattern) if pattern else None

    # Gather matching files
    all_files = []
    for fname in os.listdir(directory):
        full_path = os.path.join(directory, fname)
        if not os.path.isfile(full_path):
            continue
        if not fname.lower().endswith('.xlsx'):
            continue
        if rx:
            match = rx.match(fname)
            if match:
                all_files.append((full_path, match))
        else:
            # no pattern => accept all .xlsx
            all_files.append((full_path, None))

    if not all_files:
        raise ValueError(f"No Excel files matched in directory '{directory}' with pattern '{pattern}'.")

    # We'll parse each file => produce:
    #   (header_for_file, [dictRows], colSamples)
    # Then unify the column sets and produce final DDL + inserts.
    master_col_samples = {}  # colName => sampleVal
    file_data = []           # each entry => list of rowDict
    all_colnames_seen = set()  # to unify columns across files

    # We'll define a "master order" of columns from the FIRST file that yields data
    master_column_order = []  # we will union further columns as we go

    first_data_file_used = False

    for full_path, fmatch in all_files:
        # Extract the capturing group if any
        extra_val = None
        if filename_col:
            if fmatch and fmatch.lastindex == 1:
                extra_val = fmatch.group(1)
            else:
                # fallback => entire filename
                extra_val = os.path.basename(full_path)

        header_for_file, dict_rows, col_samples_dict = read_excel_content(
            full_path,
            config,
            extra_col_name=filename_col,
            extra_col_value=extra_val
        )
        print(dict_rows)
        if header_for_file is None or dict_rows is None:
            # means skip
            continue

        # Now unify columns
        # If it's the first data-producing file, define the initial master_column_order
        if not first_data_file_used:
            master_column_order = header_for_file[:]  # copy
            for c in master_column_order:
                all_colnames_seen.add(c)
            first_data_file_used = True
        else:
            # Add any new columns from this file that are not in master_column_order
            for c in header_for_file:
                if c not in all_colnames_seen:
                    master_column_order.append(c)
                    all_colnames_seen.add(c)

        # Merge col_samples
        if approach == "auto" and col_samples_dict:
            for c_name, c_val in col_samples_dict.items():
                if c_name not in master_col_samples and c_val is not None:
                    master_col_samples[c_name] = c_val
                    print(c_name+":"+c_val)

        file_data.append(dict_rows)

    if not first_data_file_used:
        return f"-- No valid Excel data found in directory '{directory}'", []

    # Now we have a union of all columns in master_column_order.
    # We produce final columns_info => (colName, sampleVal)
    columns_info = []
    for col_name in master_column_order:
        if approach == "auto":
            sample_val = master_col_samples.get(col_name, None)
        else:
            sample_val = None
        columns_info.append((col_name, sample_val))

    ddl = generate_ddl(table_name, columns_info, approach)

    # Build inserts from file_data
    insert_statements = []
    # Flatten file_data => array of arrays of dicts
    for dict_rows in file_data:
        for row_dict in dict_rows:
            col_list = []
            val_list = []
            for (col_name, _) in columns_info:
                raw_val = row_dict.get(col_name, None)  # Missing => None
                val_str = convert_value(raw_val, approach)
                col_list.append(f'"{col_name}"')
                val_list.append(val_str)
            ins_sql = f'INSERT INTO "{table_name}" ({", ".join(col_list)}) VALUES ({", ".join(val_list)});'
            insert_statements.append(ins_sql)

    return ddl, insert_statements

def main():
    if len(sys.argv) < 3:
        print("Usage: python excel_to_postgres.py <excel_file> <config_file> [output_sql]")
        print("Note: <excel_file> can be ignored if config['directory'] is used.")
        sys.exit(1)

    excel_file = sys.argv[1]
    config_file = sys.argv[2]
    output_sql = sys.argv[3] if len(sys.argv) > 3 else None

    config = load_config(config_file)
    table_name = config["tableName"]

    # Decide single-file or multi-file
    if "directory" in config:
        ddl, inserts = multi_excel_to_sql_inserts(config)
    else:
        # single-file usage
        ddl, inserts = single_excel_to_sql_inserts(excel_file, config)

    all_sql = [ddl] + inserts

    if output_sql:
        with open(output_sql, "w", encoding="utf-8") as f:
            f.write("\n".join(all_sql))
        print(f"SQL script saved to {output_sql}")
    else:
        print("\n".join(all_sql))

if __name__ == "__main__":
    main()
