import openpyxl
import json
import sys
import re
import os
import hashlib
import sqlite3
from datetime import datetime
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell


MAX_IDENTIFIER_LENGTH = 63  # This was for Postgres, but we'll still use it for safe column naming in SQLite

def finalize_column_name(raw_name, used_names):
    """
    Convert 'raw_name' into a safe, unique identifier:
      1) sanitize => underscores
      2) if >63 chars => short-hash approach
      3) if still collision => append _2, _3, etc., re-check length
      4) return final name, add to used_names
    """
    import re
    name = raw_name.lower()
    # Replace any non-alphanumeric with _
    name = re.sub(r'\W+', '_', name)
    # Strip leading/trailing underscores
    name = name.strip('_')
    if not name:
        name = "col"

    # If name > 63 chars => shorten with hash
    if len(name) > MAX_IDENTIFIER_LENGTH:
        name = shorten_with_hash(name)

    # Deduplicate if already used
    base = name
    counter = 2
    while name in used_names:
        candidate = f"{base}_{counter}"
        counter += 1
        if len(candidate) > MAX_IDENTIFIER_LENGTH:
            candidate = shorten_with_hash(candidate)
        name = candidate

    used_names.add(name)
    return name

def shorten_with_hash(long_name):
    """
    Keep first 50 chars + '_' + 8-char md5 hash
    Then if still >63, slice to 63.
    """
    import hashlib
    shortened = long_name[:50]
    h = hashlib.md5(long_name.encode('utf-8')).hexdigest()[:8]
    new_name = f"{shortened}_{h}"
    if len(new_name) > MAX_IDENTIFIER_LENGTH:
        new_name = new_name[:MAX_IDENTIFIER_LENGTH]
    return new_name

def load_config(config_path):
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def parse_region(region_str):
    """
    region_str like 'B2:F100'
    parse -> (start_col, start_row, end_col, end_row) zero-based
    if invalid or None => return None to read entire used range
    """
    if not region_str:
        return None

    pattern = re.compile(r'^([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)$')
    match = pattern.match(region_str)
    if not match:
        print(f"Warning: region '{region_str}' not recognized. Using full sheet.")
        return None

    start_col_letters, start_row_num, end_col_letters, end_row_num = match.groups()
    start_col = letters_to_index(start_col_letters)
    start_row = int(start_row_num) - 1
    end_col = letters_to_index(end_col_letters)
    end_row = int(end_row_num) - 1
    return (start_col, start_row, end_col, end_row)

def letters_to_index(col_letters):
    """ e.g. 'A' -> 0, 'B'->1, 'AA'->26, etc. """
    col_letters = col_letters.upper()
    result = 0
    for c in col_letters:
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result - 1

from openpyxl.utils import range_boundaries
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

def unmerge_worksheet(ws: Worksheet):
    """
    For every merged cell range, parse the range string via range_boundaries()
    to get consistent numeric row/col indices, unmerge it, and fill all cells
    with the top-left value. Logs debug info and raises on error.
    """
    # Snapshot of merged ranges to avoid modifying while iterating
    merged_ranges = list(ws.merged_cells.ranges)

    for merged_range in merged_ranges:
        range_str = merged_range.coord  # e.g. 'P17:P19' from openpyxl

        print(f"\nProcessing merged range '{range_str}'")

        # Instead of merged_range.bounds, parse the range string:
        min_colA1, min_rowA1, max_colA1, max_rowA1 = range_boundaries(range_str)
        print(f"  => range_boundaries gave: cols=({min_colA1}..{max_colA1}), rows=({min_rowA1}..{max_rowA1})")

        # If needed, swap to ensure min <= max
        if min_colA1 > max_colA1:
            min_colA1, max_colA1 = max_colA1, min_colA1
        if min_rowA1 > max_rowA1:
            min_rowA1, max_rowA1 = max_rowA1, min_rowA1

        # Now reorder them into the normal openpyxl "row, col" approach
        min_row, min_col, max_row, max_col = min_rowA1, min_colA1, max_rowA1, max_colA1
        print(f"  => Interpreted as (min_row={min_row}, min_col={min_col}, max_row={max_row}, max_col={max_col})")

        # Grab the top-left cell's value
        top_left_cell = ws.cell(row=min_row, column=min_col)
        try:
            top_left_value = top_left_cell.value
            print(f"  => top-left cell {top_left_cell.coordinate} has value {top_left_value!r}")
        except Exception as e:
            print(f"** Error reading cell {top_left_cell.coordinate}: {e}")
            continue

        # Unmerge using the original coordinate string
        try:
            ws.unmerge_cells(range_str)
        except Exception as e:
            print(f"** Error unmerging '{range_str}': {e}")
            continue

        # Fill each formerly merged cell with top_left_value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell_obj = ws.cell(row=r, column=c)
                cell_coord = cell_obj.coordinate

                if isinstance(cell_obj, MergedCell):
                    print(f"** Warning: {cell_coord} is still MergedCell somehow?!")

                try:
                    cell_obj.value = top_left_value
                except Exception as e:
                    print(f"** Error writing value to {cell_coord}: {e}")
                    raise




#
# For SQLite, let's define a simple type-guessing approach:
#   If approach='stringAll' => always TEXT
#   If approach='auto':
#       - int => INTEGER
#       - float => REAL
#       - otherwise => TEXT
#
def guess_sqlite_type(sample_val):
    if sample_val is None:
        return "TEXT"
    if isinstance(sample_val, int):
        return "INTEGER"
    if isinstance(sample_val, float):
        return "REAL"
    # You could do more robust checks for date, bool, etc.
    return "TEXT"

def convert_sqlite_value(raw_val, approach):
    """
    Convert the cell value to a Python object suitable for sqlite3's parameter binding.
    We won't manually escape—just use parameter placeholders.
    """
    if raw_val is None:
        return None

    if approach == "stringAll":
        return str(raw_val)

    # approach == 'auto'
    if isinstance(raw_val, int):
        return raw_val
    if isinstance(raw_val, float):
        return raw_val
    if isinstance(raw_val, datetime):
        # store as string or ISO format
        return raw_val.isoformat()

    # try to parse as float
    text_val = str(raw_val)
    try:
        float_val = float(text_val)
        # Distinguish between int/float if you like:
        if float_val.is_integer():
            return int(float_val)
        else:
            return float_val
    except:
        pass
    return text_val

def read_excel_content(excel_file, config, extra_col_name=None, extra_col_value=None):
    """
    Loads the Excel file, unmerges, reads the specified sheet, identifies headers,
    returns (final_header_for_file, all_dicts, col_samples).
    """
    wb = openpyxl.load_workbook(excel_file, data_only=True, keep_links=False)
    ws_name = config.get("sheetName", "Sheet1")
    if ws_name not in wb.sheetnames:
        print(f"Warning: File '{excel_file}' has no sheet '{ws_name}'; skipping.")
        return None, None, None

    ws = wb[ws_name]

    # unmerge all cells so each has correct value
    unmerge_worksheet(ws)

    region_tuple = parse_region(config.get("region"))
    approach = config.get("typeApproach", "stringAll")
    header_rows = config.get("headerRows", 1)
    header_joiner = config.get("headerJoiner", " - ")

    # figure out row/col bounds
    if region_tuple:
        sc, sr, ec, er = region_tuple
    else:
        sc = ws.min_column - 1
        sr = ws.min_row - 1
        ec = ws.max_column - 1
        er = ws.max_row - 1

    # read all rows in that region
    all_rows = []
    for row_idx in range(sr, er + 1):
        row_cells = []
        for col_idx in range(sc, ec + 1):
            cell_val = ws.cell(row=row_idx + 1, column=col_idx + 1).value
            row_cells.append(cell_val)
        all_rows.append(row_cells)

    if len(all_rows) < header_rows:
        print(f"Warning: Not enough rows for 'headerRows' in file {excel_file}, skipping.")
        return None, None, None

    # Build combined header from the top 'headerRows'
    num_cols = len(all_rows[0])
    header_labels = []
    for col_i in range(num_cols):
        parts = []
        for hr_i in range(header_rows):
            val = all_rows[hr_i][col_i]
            if val is None or str(val).strip() == "":
                val = f"col{hr_i+1}_c{col_i+1}"
            else:
                val = str(val).strip()
            parts.append(val)
        combined_name = header_joiner.join(parts)
        header_labels.append(combined_name)

    # Deduplicate/truncate them => produce final col names for THIS file
    used = set()
    final_header_for_file = []
    for raw_col_name in header_labels:
        safe_name = finalize_column_name(raw_col_name, used)
        final_header_for_file.append(safe_name)

    # Data rows => skip the 'headerRows'
    data_part = all_rows[header_rows:]

    # Convert data_part to a list of dictionaries: {colName: cellValue}
    all_dicts = []
    for row in data_part:
        row_dict = {}
        for c_idx, raw_val in enumerate(row):
            col_name = final_header_for_file[c_idx]
            row_dict[col_name] = raw_val
        # If extra col was requested, add it
        if extra_col_name:
            if extra_col_name not in final_header_for_file:
                final_header_for_file.append(extra_col_name)
            row_dict[extra_col_name] = extra_col_value
        all_dicts.append(row_dict)

    # For auto approach => find sample for each column in this file
    col_samples = {}
    if approach == "auto":
        for row_dict in all_dicts:
            for col_name, val in row_dict.items():
                if col_name not in col_samples and val is not None:
                    col_samples[col_name] = val

    return final_header_for_file, all_dicts, col_samples

def create_sqlite_table(conn, table_name, columns_info, drop_existing=True):
    """
    Creates a table in SQLite given the column names and a sample for type inference.
    columns_info is list of (col_name, sample_val).
    """
    if drop_existing:
        conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')

    # Build CREATE TABLE statement with inferred SQLite types
    col_defs = []
    for col_name, sample_val in columns_info:
        col_type = guess_sqlite_type(sample_val)
        col_defs.append(f'"{col_name}" {col_type}')

    create_stmt = f'CREATE TABLE "{table_name}" (\n  {", ".join(col_defs)}\n)'
    conn.execute(create_stmt)

def insert_into_sqlite(conn, table_name, columns_info, row_dicts, approach):
    """
    Insert the given data (list of dicts) into the SQLite table.
    columns_info: list of (col_name, sample_val)
    row_dicts: list of {col_name -> value}
    """
    col_names = [col for (col, _) in columns_info]
    placeholders = ", ".join(["?"] * len(col_names))
    col_list_str = ", ".join(f'"{cn}"' for cn in col_names)
    insert_sql = f'INSERT INTO "{table_name}" ({col_list_str}) VALUES ({placeholders})'

    for row_dict in row_dicts:
        row_values = []
        for col_name, _sample in columns_info:
            raw_val = row_dict.get(col_name)
            val_for_sqlite = convert_sqlite_value(raw_val, approach)
            row_values.append(val_for_sqlite)
        conn.execute(insert_sql, row_values)

def single_excel_to_sqlite(excel_file, config, conn):
    """
    Processes a single Excel file -> create table in SQLite, do inserts.
    """
    approach = config.get("typeApproach", "stringAll")
    table_name = config["tableName"]

    final_header_for_file, all_dicts, col_samples_dict = read_excel_content(excel_file, config)
    if final_header_for_file is None:
        print(f"-- Skipped file {excel_file}, no sheet or not enough rows.")
        return

    if col_samples_dict is None:
        col_samples_dict = {}

    # Build columns_info
    columns_info = []
    for col_name in final_header_for_file:
        sample_val = col_samples_dict.get(col_name, None) if (approach == "auto") else None
        columns_info.append((col_name, sample_val))

    # Create table & insert
    create_sqlite_table(conn, table_name, columns_info, drop_existing=True)
    insert_into_sqlite(conn, table_name, columns_info, all_dicts, approach)

def multi_excel_to_sqlite(config, conn):
    """
    Scans a directory for Excel files, merges them into one table in SQLite.
    If columns differ across files, we unify them (like your old code).
    """
    directory = config["directory"]
    pattern = config.get("filenamePattern")
    filename_col = config.get("filenameColumnName")
    approach = config.get("typeApproach", "stringAll")
    table_name = config["tableName"]

    if not os.path.isdir(directory):
        raise ValueError(f"Directory '{directory}' not found or not a directory.")

    rx = re.compile(pattern) if pattern else None

    # Gather matching files
    all_files = []
    for fname in os.listdir(directory):
        full_path = os.path.join(directory, fname)
        if not os.path.isfile(full_path):
            continue
        if not fname.lower().endswith('.xlsx'):
            continue
        if rx:
            match = rx.match(fname)
            if match:
                all_files.append((full_path, match))
        else:
            # no pattern => accept all .xlsx
            all_files.append((full_path, None))

    if not all_files:
        raise ValueError(f"No .xlsx files matched in '{directory}' with pattern '{pattern}'.")

    # Prepare to unify columns
    master_col_samples = {}
    file_data = []  # each entry => list of rowDict
    all_colnames_seen = set()
    master_column_order = []
    first_data_file_used = False

    for full_path, fmatch in all_files:
        if os.path.getsize(full_path) == 0:
            print(f"Skipping empty file: {full_path}")
            continue
    
        print(f"-- Processing file '{full_path}'.")
        # If there's a capturing group for filename, store that in the row if desired
        extra_val = None
        if filename_col:
            # If the pattern had exactly 1 capturing group
            if fmatch and fmatch.lastindex == 1:
                extra_val = fmatch.group(1)
            else:
                # fallback => entire filename
                extra_val = os.path.basename(full_path)

        header_for_file, dict_rows, col_samples_dict = read_excel_content(
            full_path,
            config,
            extra_col_name=filename_col,
            extra_col_value=extra_val
        )
        if header_for_file is None or dict_rows is None:
            # means skip
            continue

        # unify columns
        if not first_data_file_used:
            master_column_order = header_for_file[:]
            for c in master_column_order:
                all_colnames_seen.add(c)
            first_data_file_used = True
        else:
            for c in header_for_file:
                if c not in all_colnames_seen:
                    master_column_order.append(c)
                    all_colnames_seen.add(c)

        # gather sample values
        if approach == "auto" and col_samples_dict:
            for c_name, c_val in col_samples_dict.items():
                if c_name not in master_col_samples and c_val is not None:
                    master_col_samples[c_name] = c_val

        file_data.append(dict_rows)

    if not first_data_file_used:
        print(f"-- No valid Excel data found in directory '{directory}'.")
        return

    # Build final columns_info
    columns_info = []
    for col_name in master_column_order:
        if approach == "auto":
            sample_val = master_col_samples.get(col_name, None)
        else:
            sample_val = None
        columns_info.append((col_name, sample_val))

    # Create the unified table
    create_sqlite_table(conn, table_name, columns_info, drop_existing=True)

    # Insert all data
    for dict_rows in file_data:
        insert_into_sqlite(conn, table_name, columns_info, dict_rows, approach)

def main():
    if len(sys.argv) < 3:
        print("Usage: python excel_to_sqlite.py <excel_file> <config_file> [sqlite_db_path]")
        sys.exit(1)

    excel_file = sys.argv[1]
    config_file = sys.argv[2]
    # Optionally, let user specify DB path; default "output.db"
    db_path = sys.argv[3] if len(sys.argv) > 3 else "output.db"

    config = load_config(config_file)
    table_name = config["tableName"]

    # Connect to SQLite
    conn = sqlite3.connect(db_path)

    # Single-file vs. multi-file logic
    if "directory" in config:
        multi_excel_to_sqlite(config, conn)
    else:
        single_excel_to_sqlite(excel_file, config, conn)

    conn.commit()
    conn.close()
    print(f"Done. Data loaded into SQLite database: {db_path}")

if __name__ == "__main__":
    main()
